import os
from datetime import datetime
from openpyxl import load_workbook
from django.conf import settings

from personal.models import Trabajador
from evaluaciones.models import Intento
from entrenamientos.models import AvanceEntrenamiento
from supervision.models import RevisionSupervisor
from autorizaciones.models import AutorizacionFinal


def escribir_celda(ws, fila, col, valor):
    """Escribe en celda normal o maestra de rango combinado."""
    celda = ws.cell(row=fila, column=col)
    coord = celda.coordinate
    for rango in ws.merged_cells.ranges:
        if coord in rango:
            maestra = ws.cell(row=rango.min_row, column=rango.min_col)
            try:
                maestra.value = valor
            except AttributeError:
                pass
            return
    try:
        celda.value = valor
    except AttributeError:
        pass


def llenar_hoja_induccion(ws, trabajador, intentos_gestion, intentos_tecnicos, revisiones):
    """
    Coordenadas exactas del F-52 real de LABSAF:

    DATOS PERSONALES:
      E10 = Nombre y Apellidos
      E11 = Cargo
      E12 = Laboratorio / área
      E13 = Fecha de Ingreso
      E14 = Fecha de Inicio de Entrenamiento

    SECCIÓN 2 — Sistema de Gestión (filas 21-27):
      Col F(6)=Fecha 1ra  Col G(7)=Nota 1ra
      Col H(8)=Fecha 2da  Col I(9)=Nota 2da
      Col J(10)=Fecha 3ra Col K(11)=Nota 3ra
      Col L(12)=Fecha Satisfactorio  Col M(13)=Responsable

    SECCIÓN 3 — Técnico General (filas 35-41):
      Col E(5)=Fecha Eval1  Col F(6)=Nota Eval1
      Col G(7)=Fecha Eval2  Col H(8)=Nota Eval2
      Col I(9)=Fecha Sup1   Col J(10)=Conclusión Sup1
      Col K(11)=Fecha Sup2  Col L(12)=Conclusión Sup2
      Col M(13)=Fecha Satisfactorio  Col N(14)=Responsable
    """

    # ── Datos del personal ────────────────────────────────────────────────────
    escribir_celda(ws, 10, 5, trabajador.usuario.get_full_name())
    escribir_celda(ws, 11, 5, str(trabajador.cargo) if trabajador.cargo else '')
    escribir_celda(ws, 12, 5, str(trabajador.area)  if trabajador.area  else '')
    if trabajador.fecha_ingreso:
        escribir_celda(ws, 13, 5, trabajador.fecha_ingreso)
        escribir_celda(ws, 14, 5, trabajador.fecha_ingreso)

    # Supervisor para el campo "Responsable"
    rev1 = revisiones[0] if len(revisiones) > 0 else None
    rev2 = revisiones[1] if len(revisiones) > 1 else None
    nombre_resp = rev1.supervisor.get_full_name() if rev1 and rev1.supervisor else ''

    # ── SECCIÓN 2: Sistema de Gestión ────────────────────────────────────────
    FILAS_GESTION = [21, 22, 23, 24, 25, 26, 27]

    # Agrupar intentos por evaluación manteniendo orden
    evals_gestion = []
    visto = set()
    for intento in intentos_gestion:
        if intento.evaluacion_id not in visto:
            visto.add(intento.evaluacion_id)
            evals_gestion.append(intento.evaluacion_id)

    for idx, eval_id in enumerate(evals_gestion):
        if idx >= len(FILAS_GESTION):
            break
        fila = FILAS_GESTION[idx]
        intentos_ev = sorted(
            [i for i in intentos_gestion if i.evaluacion_id == eval_id],
            key=lambda x: x.numero_intento
        )

        if len(intentos_ev) >= 1:
            i1 = intentos_ev[0]
            escribir_celda(ws, fila, 6, i1.fecha_inicio.date() if i1.fecha_inicio else None)
            escribir_celda(ws, fila, 7, float(i1.puntuacion))

        if len(intentos_ev) >= 2:
            i2 = intentos_ev[1]
            escribir_celda(ws, fila, 8, i2.fecha_inicio.date() if i2.fecha_inicio else None)
            escribir_celda(ws, fila, 9, float(i2.puntuacion))

        if len(intentos_ev) >= 3:
            i3 = intentos_ev[2]
            escribir_celda(ws, fila, 10, i3.fecha_inicio.date() if i3.fecha_inicio else None)
            escribir_celda(ws, fila, 11, float(i3.puntuacion))

        aprobado = next((i for i in intentos_ev if i.aprobado), None)
        if aprobado:
            escribir_celda(ws, fila, 12, aprobado.fecha_fin.date() if aprobado.fecha_fin else None)
            if nombre_resp:
                escribir_celda(ws, fila, 13, nombre_resp)

    # ── SECCIÓN 3: Técnico General ────────────────────────────────────────────
    FILAS_TECNICO = [35, 36, 37, 38, 39, 40, 41]

    evals_tecnico = []
    visto = set()
    for intento in intentos_tecnicos:
        if intento.evaluacion_id not in visto:
            visto.add(intento.evaluacion_id)
            evals_tecnico.append(intento.evaluacion_id)

    for idx, eval_id in enumerate(evals_tecnico):
        if idx >= len(FILAS_TECNICO):
            break
        fila = FILAS_TECNICO[idx]
        intentos_ev = sorted(
            [i for i in intentos_tecnicos if i.evaluacion_id == eval_id],
            key=lambda x: x.numero_intento
        )

        if len(intentos_ev) >= 1:
            i1 = intentos_ev[0]
            escribir_celda(ws, fila, 5, i1.fecha_inicio.date() if i1.fecha_inicio else None)
            escribir_celda(ws, fila, 6, float(i1.puntuacion))

        if len(intentos_ev) >= 2:
            i2 = intentos_ev[1]
            escribir_celda(ws, fila, 7, i2.fecha_inicio.date() if i2.fecha_inicio else None)
            escribir_celda(ws, fila, 8, float(i2.puntuacion))

        if rev1:
            escribir_celda(ws, fila, 9,  rev1.fecha.date() if rev1.fecha else None)
            escribir_celda(ws, fila, 10, 'Satisfactorio' if rev1.estado == 'aprobado' else rev1.get_estado_display())

        if rev2:
            escribir_celda(ws, fila, 11, rev2.fecha.date() if rev2.fecha else None)
            escribir_celda(ws, fila, 12, 'Satisfactorio' if rev2.estado == 'aprobado' else rev2.get_estado_display())

        aprobado = next((i for i in intentos_ev if i.aprobado), None)
        if aprobado:
            escribir_celda(ws, fila, 13, aprobado.fecha_fin.date() if aprobado.fecha_fin else None)
        if nombre_resp:
            escribir_celda(ws, fila, 14, nombre_resp)

    # Firma supervisor pie de página (fila 49)
    if nombre_resp:
        escribir_celda(ws, 49, 7, nombre_resp)
    if rev1 and rev1.fecha:
        escribir_celda(ws, 49, 9, rev1.fecha.date())


def llenar_hoja_reverso(ws, trabajador, avance, revisiones):
    """Llena hoja de Entrenamiento Técnico Específico."""
    rev1 = revisiones[0] if len(revisiones) > 0 else None
    rev2 = revisiones[1] if len(revisiones) > 1 else None

    if avance and avance.fecha_inicio:
        fecha_ini = avance.fecha_inicio.date()
        fecha_fin = avance.fecha_completado.date() if avance.fecha_completado else None
        texto = f"{fecha_ini.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}" if fecha_fin else fecha_ini.strftime('%d/%m/%Y')
        escribir_celda(ws, 6, 5, texto)
        escribir_celda(ws, 6, 12, fecha_ini)
        if fecha_fin:
            escribir_celda(ws, 6, 16, fecha_fin)

    if avance and avance.modulo and avance.modulo.evaluacion:
        intentos = list(
            Intento.objects.filter(
                trabajador=trabajador,
                evaluacion=avance.modulo.evaluacion
            ).order_by('numero_intento')
        )
        for i, fila in enumerate(range(12, 18)):
            intento = intentos[i] if i < len(intentos) else None
            if intento:
                escribir_celda(ws, fila, 5, intento.fecha_inicio.date() if intento.fecha_inicio else None)
                escribir_celda(ws, fila, 6, float(intento.puntuacion))
            if rev1:
                escribir_celda(ws, fila, 10, rev1.fecha.date() if rev1.fecha else None)
                escribir_celda(ws, fila, 11, 'Satisfactorio' if rev1.estado == 'aprobado' else rev1.get_estado_display())
            if rev2:
                escribir_celda(ws, fila, 12, rev2.fecha.date() if rev2.fecha else None)
                escribir_celda(ws, fila, 13, 'Satisfactorio' if rev2.estado == 'aprobado' else rev2.get_estado_display())

    try:
        aut = AutorizacionFinal.objects.get(trabajador=trabajador)
        if aut.estado == 'autorizado' and aut.fecha_resolucion:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and 'Autorización Para' in str(cell.value):
                        escribir_celda(ws, cell.row, 5, 'Autorización Total')
                        escribir_celda(ws, cell.row, 12, aut.fecha_resolucion.date())
                        break
    except AutorizacionFinal.DoesNotExist:
        pass


def generar_excel_trabajador(trabajador_id: int) -> str:
    trabajador = Trabajador.objects.select_related('usuario', 'cargo', 'area').get(pk=trabajador_id)

    ruta_plantilla = os.path.join(settings.MEDIA_ROOT, 'plantillas', 'F-52_plantilla_base.xlsx')
    if not os.path.exists(ruta_plantilla):
        raise FileNotFoundError(f'Plantilla no encontrada: {ruta_plantilla}')

    wb = load_workbook(ruta_plantilla)

    # Obtener todos los intentos
    todos_intentos = list(
        Intento.objects.filter(trabajador=trabajador)
        .select_related('evaluacion')
        .order_by('evaluacion__fecha_creacion', 'numero_intento')
    )

    # Separar por tipo de evaluación
    intentos_gestion  = [i for i in todos_intentos if i.evaluacion.tipo == 'gestion']
    intentos_tecnicos = [i for i in todos_intentos if i.evaluacion.tipo == 'tecnico']

    # Si están clasificados como 'induccion', separar por nombre
    if not intentos_gestion and not intentos_tecnicos:
        for intento in todos_intentos:
            titulo = intento.evaluacion.titulo.lower()
            if any(p in titulo for p in ['gestion', 'gestión', 'sistema', 'management']):
                intentos_gestion.append(intento)
            else:
                intentos_tecnicos.append(intento)

    revisiones = list(
        RevisionSupervisor.objects.filter(trabajador=trabajador)
        .select_related('supervisor').order_by('fecha')
    )

    avances = list(
        AvanceEntrenamiento.objects.filter(trabajador=trabajador)
        .select_related('modulo', 'modulo__evaluacion')
        .order_by('modulo__tipo', 'fecha_inicio')
    )

    # Llenar hoja principal
    if 'Inducción' in wb.sheetnames:
        llenar_hoja_induccion(wb['Inducción'], trabajador, intentos_gestion, intentos_tecnicos, revisiones)

    # Llenar hojas reverso
    for nombre_hoja in [n for n in wb.sheetnames if n.startswith('Reverso-')]:
        ws = wb[nombre_hoja]
        avance = next(
            (a for a in avances if a.modulo.titulo[:15].lower() in nombre_hoja.lower()),
            avances[0] if avances else None
        )
        llenar_hoja_reverso(ws, trabajador, avance, revisiones)

    # Matriz de seguimiento
    if 'Matriz de Seguimiento' in wb.sheetnames:
        escribir_celda(wb['Matriz de Seguimiento'], 2, 1,
            f"Elaborado para: {trabajador.usuario.get_full_name()} — {datetime.now().strftime('%d/%m/%Y')}"
        )

    # Guardar
    nombre = f"F-52_{trabajador.usuario.last_name.replace(' ','_')}_{trabajador.usuario.first_name.replace(' ','_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    carpeta = os.path.join(settings.MEDIA_ROOT, 'reportes')
    os.makedirs(carpeta, exist_ok=True)
    ruta_salida = os.path.join(carpeta, nombre)
    wb.save(ruta_salida)
    return os.path.join('reportes', nombre)